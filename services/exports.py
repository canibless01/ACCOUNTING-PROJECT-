"""
Export service — generates PDF and Excel reports from query results.
Supports P&L, Cash Flow, and Reconciliation Summary reports.
"""
from __future__ import annotations

import io
import os
from datetime import date, datetime
from typing import Any

from config import Config

# Ensure export temp directory exists
os.makedirs(Config.EXPORT_TEMP_DIR, exist_ok=True)


# ── PDF export ─────────────────────────────────────────────────────────────────

def generate_pdf_report(report_type: str, data: dict, user_info: dict) -> bytes:
    """
    Generate a PDF report and return raw bytes.
    report_type: 'pl' | 'cashflow' | 'reconciliation'
    """
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle
    )

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=2 * cm,
        leftMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "Title", parent=styles["Heading1"], fontSize=18, spaceAfter=6
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"], fontSize=11, textColor=colors.grey, spaceAfter=20
    )
    header_style = ParagraphStyle(
        "SectionHeader", parent=styles["Heading2"], fontSize=13, spaceAfter=8
    )

    elements = []

    # ── Header ────────────────────────────────────────────────────────────────
    report_titles = {
        "pl": "Profit & Loss Statement",
        "cashflow": "Cash Flow Statement",
        "reconciliation": "Reconciliation Summary",
    }
    elements.append(Paragraph("Holy Grills", title_style))
    elements.append(Paragraph(report_titles.get(report_type, "Financial Report"), header_style))

    date_range = data.get("date_range", {})
    if date_range:
        elements.append(
            Paragraph(
                f"Period: {date_range.get('from', '')} to {date_range.get('to', '')}",
                subtitle_style,
            )
        )
    elements.append(
        Paragraph(f"Generated: {datetime.now().strftime('%d %B %Y, %H:%M')}", subtitle_style)
    )
    elements.append(Spacer(1, 0.5 * cm))

    # ── Report body ───────────────────────────────────────────────────────────
    if report_type == "pl":
        elements.extend(_build_pl_pdf_elements(data, styles, colors))
    elif report_type == "cashflow":
        elements.extend(_build_cashflow_pdf_elements(data, styles, colors))
    elif report_type == "reconciliation":
        elements.extend(_build_reconciliation_pdf_elements(data, styles, colors))

    doc.build(elements)
    buffer.seek(0)
    return buffer.read()


def _build_pl_pdf_elements(data: dict, styles, colors) -> list:
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import Paragraph, Spacer, Table, TableStyle
    from reportlab.lib.units import cm

    elements = []
    currency = "₦"

    # Income section
    elements.append(Paragraph("Income", styles["Heading2"]))
    income_rows = data.get("income_by_category", [])
    if income_rows:
        table_data = [["Category", "Amount"]]
        for row in income_rows:
            table_data.append([
                row.get("category_name", "Uncategorized"),
                f"{currency}{float(row.get('total', 0)):,.2f}",
            ])
        total_income = data.get("total_income", 0)
        table_data.append(["Total Income", f"{currency}{float(total_income):,.2f}"])
        t = Table(table_data, colWidths=[12 * cm, 5 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#10B981")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9FAFB")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#D1FAE5")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
    elements.append(Spacer(1, 0.5 * cm))

    # Expense section
    elements.append(Paragraph("Expenses", styles["Heading2"]))
    expense_rows = data.get("expense_by_category", [])
    if expense_rows:
        table_data = [["Category", "Amount"]]
        for row in expense_rows:
            table_data.append([
                row.get("category_name", "Uncategorized"),
                f"{currency}{float(row.get('total', 0)):,.2f}",
            ])
        total_expense = data.get("total_expense", 0)
        table_data.append(["Total Expenses", f"{currency}{float(total_expense):,.2f}"])
        t = Table(table_data, colWidths=[12 * cm, 5 * cm])
        t.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#EF4444")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#F9FAFB")]),
            ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FEE2E2")),
            ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
            ("ALIGN", (1, 0), (1, -1), "RIGHT"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
            ("TOPPADDING", (0, 0), (-1, -1), 6),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ]))
        elements.append(t)
    elements.append(Spacer(1, 0.5 * cm))

    # Net profit
    net = float(data.get("net_profit", 0))
    net_color = "#10B981" if net >= 0 else "#EF4444"
    net_label = "Net Profit" if net >= 0 else "Net Loss"
    summary_data = [[net_label, f"{currency}{abs(net):,.2f}"]]
    t = Table(summary_data, colWidths=[12 * cm, 5 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor(net_color)),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.white),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 14),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("TOPPADDING", (0, 0), (-1, -1), 10),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
    ]))
    elements.append(t)
    elements.append(Paragraph("Note: Internal transfers excluded from this statement.", styles["Italic"]))
    return elements


def _build_cashflow_pdf_elements(data: dict, styles, colors) -> list:
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import cm

    currency = "₦"
    elements = []
    rows = [
        ["", ""],
        ["Opening Balance", f"{currency}{float(data.get('opening_balance', 0)):,.2f}"],
        ["Total Inflows (Credit)", f"{currency}{float(data.get('total_in', 0)):,.2f}"],
        ["Total Outflows (Debit)", f"({currency}{float(data.get('total_out', 0)):,.2f})"],
        ["Calculated Closing Balance", f"{currency}{float(data.get('closing_balance', 0)):,.2f}"],
        ["Actual Closing Balance", f"{currency}{float(data.get('actual_closing', 0)):,.2f}" if data.get('actual_closing') else "N/A"],
    ]
    t = Table(rows, colWidths=[12 * cm, 5 * cm])
    t.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -2), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, 3), (-1, 3), colors.HexColor("#FEE2E2")),
        ("BACKGROUND", (0, 4), (-1, 4), colors.HexColor("#D1FAE5")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    return elements


def _build_reconciliation_pdf_elements(data: dict, styles, colors) -> list:
    from reportlab.platypus import Table, TableStyle, Paragraph
    from reportlab.lib.units import cm

    currency = "₦"
    elements = []
    mismatches = data.get("mismatches", [])
    if not mismatches:
        elements.append(Paragraph("✓ All accounts are reconciled.", styles["Normal"]))
        return elements

    rows = [["Account", "Calculated", "Actual", "Gap", "Since"]]
    for m in mismatches:
        rows.append([
            m.get("name", ""),
            f"{currency}{float(m.get('calculated_balance', 0)):,.2f}",
            f"{currency}{float(m.get('actual_balance', 0)):,.2f}" if m.get("actual_balance") else "N/A",
            f"{currency}{float(m.get('reconciliation_gap', 0)):,.2f}",
            str(m.get("reconciliation_mismatch_since", ""))[:10],
        ])
    t = Table(rows, colWidths=[5 * cm, 4 * cm, 4 * cm, 3 * cm, 3 * cm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F59E0B")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FEF3C7")]),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#E5E7EB")),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(t)
    return elements


# ── Excel export ───────────────────────────────────────────────────────────────

def generate_excel_report(report_type: str, data: dict, user_info: dict) -> bytes:
    """Generate an Excel (.xlsx) report and return raw bytes."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Alignment, Border, Font, PatternFill, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    report_titles = {
        "pl": "Profit & Loss",
        "cashflow": "Cash Flow",
        "reconciliation": "Reconciliation",
    }
    ws.title = report_titles.get(report_type, "Report")

    # Header styling helpers
    header_fill = PatternFill("solid", fgColor="1E3A5F")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    income_fill = PatternFill("solid", fgColor="D1FAE5")
    expense_fill = PatternFill("solid", fgColor="FEE2E2")
    total_font = Font(bold=True, size=11)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    def apply_header(cell, value):
        cell.value = value
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

    def apply_currency(cell, value):
        cell.value = float(value) if value is not None else 0.0
        cell.number_format = '₦#,##0.00'
        cell.border = thin_border

    row = 1
    ws.cell(row=row, column=1, value="Holy Grills").font = Font(bold=True, size=16)
    row += 1
    ws.cell(row=row, column=1, value=ws.title).font = Font(bold=True, size=13)
    row += 1
    date_range = data.get("date_range", {})
    if date_range:
        ws.cell(row=row, column=1, value=f"Period: {date_range.get('from','')} to {date_range.get('to','')}")
    row += 1
    ws.cell(row=row, column=1, value=f"Generated: {datetime.now().strftime('%d %B %Y %H:%M')}")
    row += 2

    if report_type == "pl":
        row = _build_pl_excel(ws, data, row, apply_header, apply_currency, total_font, income_fill, expense_fill, thin_border)
    elif report_type == "cashflow":
        row = _build_cashflow_excel(ws, data, row, apply_header, apply_currency, total_font, thin_border)
    elif report_type == "reconciliation":
        row = _build_recon_excel(ws, data, row, apply_header, apply_currency, total_font, thin_border)

    # Auto-fit columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.read()


def _build_pl_excel(ws, data, start_row, apply_header, apply_currency, total_font, income_fill, expense_fill, thin_border):
    from openpyxl.styles import PatternFill, Font, Border, Side

    row = start_row
    apply_header(ws.cell(row=row, column=1), "Category")
    apply_header(ws.cell(row=row, column=2), "Amount (₦)")
    row += 1

    # Income
    ws.cell(row=row, column=1, value="INCOME").font = Font(bold=True, size=12)
    row += 1
    total_income = 0.0
    for item in data.get("income_by_category", []):
        ws.cell(row=row, column=1, value=item.get("category_name", "Uncategorized"))
        apply_currency(ws.cell(row=row, column=2), item.get("total", 0))
        ws.cell(row=row, column=1).fill = income_fill
        ws.cell(row=row, column=2).fill = income_fill
        total_income += float(item.get("total", 0))
        row += 1
    ws.cell(row=row, column=1, value="Total Income").font = total_font
    apply_currency(ws.cell(row=row, column=2), total_income)
    ws.cell(row=row, column=2).font = total_font
    row += 2

    # Expenses
    ws.cell(row=row, column=1, value="EXPENSES").font = Font(bold=True, size=12)
    row += 1
    total_expense = 0.0
    for item in data.get("expense_by_category", []):
        ws.cell(row=row, column=1, value=item.get("category_name", "Uncategorized"))
        apply_currency(ws.cell(row=row, column=2), item.get("total", 0))
        ws.cell(row=row, column=1).fill = expense_fill
        ws.cell(row=row, column=2).fill = expense_fill
        total_expense += float(item.get("total", 0))
        row += 1
    ws.cell(row=row, column=1, value="Total Expenses").font = total_font
    apply_currency(ws.cell(row=row, column=2), total_expense)
    ws.cell(row=row, column=2).font = total_font
    row += 2

    net = total_income - total_expense
    ws.cell(row=row, column=1, value="NET PROFIT / LOSS").font = Font(bold=True, size=13)
    apply_currency(ws.cell(row=row, column=2), net)
    ws.cell(row=row, column=2).font = Font(bold=True, size=13)
    row += 2
    ws.cell(row=row, column=1, value="* Internal transfers excluded").font = Font(italic=True, color="888888")
    return row


def _build_cashflow_excel(ws, data, start_row, apply_header, apply_currency, total_font, thin_border):
    row = start_row
    apply_header(ws.cell(row=row, column=1), "Item")
    apply_header(ws.cell(row=row, column=2), "Amount (₦)")
    row += 1
    items = [
        ("Opening Balance", data.get("opening_balance", 0)),
        ("Total Inflows", data.get("total_in", 0)),
        ("Total Outflows", data.get("total_out", 0)),
        ("Calculated Closing Balance", data.get("closing_balance", 0)),
        ("Actual Closing Balance", data.get("actual_closing", 0) or 0),
    ]
    for label, val in items:
        ws.cell(row=row, column=1, value=label)
        apply_currency(ws.cell(row=row, column=2), val)
        row += 1
    return row


def _build_recon_excel(ws, data, start_row, apply_header, apply_currency, total_font, thin_border):
    row = start_row
    cols = ["Account", "Calculated (₦)", "Actual (₦)", "Gap (₦)", "Mismatch Since"]
    for i, col in enumerate(cols, 1):
        apply_header(ws.cell(row=row, column=i), col)
    row += 1
    for m in data.get("mismatches", []):
        ws.cell(row=row, column=1, value=m.get("name", ""))
        apply_currency(ws.cell(row=row, column=2), m.get("calculated_balance", 0))
        apply_currency(ws.cell(row=row, column=3), m.get("actual_balance", 0))
        apply_currency(ws.cell(row=row, column=4), m.get("reconciliation_gap", 0))
        ws.cell(row=row, column=5, value=str(m.get("reconciliation_mismatch_since", ""))[:10])
        row += 1
    if not data.get("mismatches"):
        ws.cell(row=row, column=1, value="All accounts are reconciled.")
    return row
